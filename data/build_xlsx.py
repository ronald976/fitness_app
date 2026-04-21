"""Transform 2026 fitness log text files into a well-formatted xlsx workbook.

Input  : 2026_log_1.txt, 2026_log_2.txt (next to this script)
Output : 2026_fitness_log.xlsx

Run    : micromamba run -n 1 python build_xlsx.py
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_GLOB = "*_log_*.txt"
DEFAULT_OUTPUT = "fitness_log.xlsx"
FILENAME_YEAR_RE = re.compile(r"(\d{4})_log")

# --- regex ------------------------------------------------------------------

DATE_RE = re.compile(r"^(\d{1,2})/(\d{1,2})\s+(.+?)\s*$")
TIME_RE = re.compile(r"^(\d{1,2}:\d{2})\s*$")
PARENS_LINE_RE = re.compile(r"^\((.+)\)\s*$")
YEAR_MARKER_RE = re.compile(r"^(20\d{2})\s*$")
# A single set token: optional weight, 'x', optional reps, optional letter tag,
# optional "+marker" suffix (e.g. +bo, +be). Examples:
#   110x7, 82x5f, 75x6+bo, x8, 65x, 26x7
SET_TOKEN_RE = re.compile(
    r"^(\d+(?:\.\d+)?)?x(\d+)?([a-z]+)?(?:\+(\w+))?$",
    re.IGNORECASE,
)
PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%")
PR_MARKERS = ("new pr", "pr match", "recent pr")

# --- data classes -----------------------------------------------------------


@dataclass
class SetEntry:
    weight: float | None
    reps: int | None
    tag: str = ""
    incomplete: bool = False


@dataclass
class Exercise:
    name: str
    sets: list[SetEntry] = field(default_factory=list)
    quick_sets: int | None = None
    intensity_pct: float | None = None
    is_pr: bool = False
    pr_text: str = ""
    notes: list[str] = field(default_factory=list)


@dataclass
class Session:
    date: date
    session_type: str
    time: str | None = None
    exercises: list[Exercise] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


# --- parsing ----------------------------------------------------------------


def _is_set_token(tok: str) -> bool:
    if "x" not in tok.lower():
        return False
    if not any(c.isdigit() for c in tok):
        return False
    return bool(SET_TOKEN_RE.match(tok))


def _parse_set_token(tok: str) -> SetEntry | None:
    m = SET_TOKEN_RE.match(tok)
    if not m:
        return None
    w_str, r_str, suffix, plus = m.groups()
    weight = float(w_str) if w_str else None
    reps = int(r_str) if r_str else None
    incomplete = reps is None
    tag_parts = []
    if suffix:
        tag_parts.append(suffix.lower())
    if plus:
        tag_parts.append(f"+{plus.lower()}")
    return SetEntry(
        weight=weight,
        reps=reps,
        tag=",".join(tag_parts),
        incomplete=incomplete,
    )


YEAR_PENALTY_DAYS = 300  # per-year distance from the file's base year


def _make_date(year: int, a: int, b: int, prev: date | None) -> date:
    """Resolve an ambiguous 'a/b' date token. The log mostly uses day/month
    but occasionally slips into month/day (e.g. "5/30 pull", "08/02 combi").
    Tries both orderings in the file's base year and the adjacent years,
    and picks the candidate that minimises distance from the previous
    session date plus a penalty for deviating from the base year. The
    base-year penalty prevents an isolated typo (e.g. "25/11" meant for
    "25/1") from snowballing every subsequent session into the next year,
    while being small enough that a real Dec→Jan rollover still wins."""
    candidates: list[date] = []
    for yr in (year - 1, year, year + 1):
        for d, m in ((a, b), (b, a)):
            try:
                dt = date(yr, m, d)
            except ValueError:
                continue
            if dt not in candidates:
                candidates.append(dt)
    if not candidates:
        raise ValueError(f"invalid date tokens: {a}/{b} (year {year})")
    if prev is None:
        try:
            return date(year, b, a)
        except ValueError:
            return candidates[0]

    def score(d: date) -> int:
        return abs((d - prev).days) + YEAR_PENALTY_DAYS * abs(d.year - year)

    return min(candidates, key=score)


def parse_log(text: str, base_year: int) -> tuple[list[Session], list[str]]:
    """Parse one log file. `base_year` is the default year for undated entries;
    explicit "YYYY" marker lines and month rollovers override it."""
    sessions: list[Session] = []
    pr_lines: list[str] = []
    lines = text.splitlines()

    # Find first date-line or year-marker line to skip email header.
    start = 0
    for i, ln in enumerate(lines):
        s = ln.strip()
        if DATE_RE.match(s) or YEAR_MARKER_RE.match(s):
            start = i
            break

    current_session: Session | None = None
    current_exercise: Exercise | None = None
    pending_time: str | None = None
    in_prs = False
    current_year = base_year
    prev_date: date | None = None

    for raw in lines[start:]:
        ln = raw.strip()
        if in_prs:
            # Preserve blank lines — they delimit PR groupings.
            pr_lines.append(raw)
            continue
        if not ln:
            continue
        if ln.startswith("PRs") or set(ln) == {"_"}:
            in_prs = True
            continue

        m = YEAR_MARKER_RE.match(ln)
        if m:
            current_year = int(m.group(1))
            prev_date = None
            continue

        m = TIME_RE.match(ln)
        if m:
            pending_time = m.group(1)
            continue

        m = DATE_RE.match(ln)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            stype = m.group(3).strip()
            dt = _make_date(current_year, a, b, prev_date)
            # Carry year forward if _make_date advanced past year boundary.
            current_year = dt.year
            prev_date = dt
            current_session = Session(
                date=dt,
                session_type=stype,
                time=pending_time,
            )
            sessions.append(current_session)
            pending_time = None
            current_exercise = None
            continue

        m = PARENS_LINE_RE.match(ln)
        if m:
            content = m.group(1).strip()
            if current_exercise is not None:
                _apply_parenthetical(current_exercise, content)
            elif current_session is not None:
                current_session.notes.append(content)
            continue

        tokens = ln.split()
        trailing_paren: str | None = None
        # Pull off trailing "(...)", e.g. "(94%)" or "(new PR!)"
        if tokens and tokens[-1].endswith(")"):
            # Collapse tokens from the first opening paren.
            for i, t in enumerate(tokens):
                if t.startswith("("):
                    trailing_paren = " ".join(tokens[i:])[1:-1].strip()
                    tokens = tokens[:i]
                    break

        if tokens and all(_is_set_token(t) for t in tokens) and current_exercise is not None:
            for t in tokens:
                s = _parse_set_token(t)
                if s is not None:
                    current_exercise.sets.append(s)
            if trailing_paren:
                _apply_parenthetical(current_exercise, trailing_paren)
            continue

        if current_session is None:
            continue

        # Exercise line. May have trailing "xN" shortcut (e.g., "Abs x3").
        # Only treat as quick-sets when N ≤ 10; higher numbers like "x60"
        # are rep counts (e.g., "Calf raises x60"), not set counts.
        name = ln
        qm = re.match(r"^(.+?)\s+x(\d+)\s*$", name)
        quick = None
        if qm:
            n = int(qm.group(2))
            if n <= 10:
                name = qm.group(1).strip()
                quick = n
        current_exercise = Exercise(name=name, quick_sets=quick)
        current_session.exercises.append(current_exercise)

    return sessions, pr_lines


def _apply_parenthetical(ex: Exercise, content: str) -> None:
    pct_match = PCT_RE.search(content)
    if pct_match:
        ex.intensity_pct = float(pct_match.group(1))
    lc = content.lower()
    if any(k in lc for k in PR_MARKERS) or ("pr" in lc and "!" in lc):
        ex.is_pr = True
        if not ex.pr_text:
            ex.pr_text = content
    # Keep the raw text as an exercise note unless it's a pure "NN%" marker.
    if not (pct_match and pct_match.group(0).strip() == content.strip()):
        ex.notes.append(content)


# --- workbook ---------------------------------------------------------------

SESSION_COLORS = {
    "pull": "DCE6F1",
    "push": "FCE4D6",
    "legs": "E2EFDA",
    "upper": "FFF2CC",
    "lower": "EDEDED",
}
PR_FILL = PatternFill("solid", fgColor="F4B183")
MIN_SESSIONS_PER_YEAR = 10
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _row_fill(session_type: str) -> PatternFill | None:
    s = session_type.lower()
    for key, color in SESSION_COLORS.items():
        if key in s:
            return PatternFill("solid", fgColor=color)
    return None


def _write_header(ws, headers: list[str]) -> None:
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            for line in v.splitlines() or [""]:
                longest = max(longest, len(line))
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), max_width)


def _write_sets_sheet(ws, sessions: list[Session]) -> None:
    headers = [
        "Session#", "Date", "Weekday", "Type", "Time",
        "Exercise", "Set#", "Weight (kg)", "Reps",
        "Tag", "Intensity %", "PR", "Notes",
    ]
    _write_header(ws, headers)

    row = 2
    for s_idx, sess in enumerate(sessions, 1):
        fill = _row_fill(sess.session_type)
        for ex in sess.exercises:
            if not ex.sets and ex.quick_sets is None:
                # Skipped or data-less exercise; still record.
                _write_row(ws, row, [
                    s_idx, sess.date, sess.date.strftime("%a"),
                    sess.session_type, sess.time,
                    ex.name, None, None, None, None,
                    ex.intensity_pct,
                    "Y" if ex.is_pr else "",
                    "; ".join(ex.notes),
                ], fill=fill, pr=ex.is_pr)
                row += 1
                continue
            if not ex.sets and ex.quick_sets is not None:
                # "xN" shortcut: emit one summary row.
                _write_row(ws, row, [
                    s_idx, sess.date, sess.date.strftime("%a"),
                    sess.session_type, sess.time,
                    ex.name, f"x{ex.quick_sets}", None, None, "quick",
                    ex.intensity_pct,
                    "Y" if ex.is_pr else "",
                    "; ".join(ex.notes),
                ], fill=fill, pr=ex.is_pr)
                row += 1
                continue
            for i, st in enumerate(ex.sets, 1):
                reps = st.reps if not st.incomplete else None
                tag = st.tag
                if st.incomplete:
                    tag = (tag + ",incomplete").strip(",")
                _write_row(ws, row, [
                    s_idx, sess.date, sess.date.strftime("%a"),
                    sess.session_type, sess.time,
                    ex.name, i, st.weight, reps, tag,
                    ex.intensity_pct,
                    "Y" if ex.is_pr else "",
                    "; ".join(ex.notes) if i == 1 else "",
                ], fill=fill, pr=ex.is_pr)
                row += 1

    for col_idx, label in enumerate(headers, 1):
        if label == "Date":
            for r in range(2, row):
                ws.cell(row=r, column=col_idx).number_format = "yyyy-mm-dd"
    _autosize(ws)


def _write_row(ws, row: int, values: list, fill=None, pr: bool = False) -> None:
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = BORDER
        if pr:
            c.fill = PR_FILL
        elif fill is not None:
            c.fill = fill
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = Alignment(horizontal="right")


def _write_sessions_sheet(ws, sessions: list[Session]) -> None:
    headers = [
        "Session#", "Date", "Weekday", "Type", "Time",
        "Exercises", "Total Sets", "PRs", "Notes",
    ]
    _write_header(ws, headers)
    for i, sess in enumerate(sessions, 1):
        fill = _row_fill(sess.session_type)
        total_sets = sum(len(e.sets) for e in sess.exercises)
        pr_names = [e.name for e in sess.exercises if e.is_pr]
        _write_row(ws, i + 1, [
            i, sess.date, sess.date.strftime("%a"),
            sess.session_type, sess.time,
            len(sess.exercises), total_sets,
            ", ".join(pr_names),
            " | ".join(sess.notes),
        ], fill=fill)
        ws.cell(row=i + 1, column=2).number_format = "yyyy-mm-dd"
    _autosize(ws)


GROUP_ORDER = ["Press", "Squat", "Deadlift", "Compound", "Non compound"]
GROUP_FILL = PatternFill("solid", fgColor="B4C7E7")
GROUP_FONT = Font(bold=True, color="1F3864")

PROG_MIN_DAYS = 10
PROG_INACTIVE_LOOKBACK = 3  # quarters

EXERCISE_ALIASES = {
    "Leg press (free weight)": "Leg press",
    "Calf raises (free weight)": "Calf raises",
    "Smith incline bench": "Incline bench smith",
}


def _normalize_exercise(name: str) -> str:
    n = EXERCISE_ALIASES.get(name, name)
    # Normalise "Smith" capitalisation drift across log entries.
    n = re.sub(r"\bSmith\b", "smith", n)
    return EXERCISE_ALIASES.get(n, n)


def _progression_group(name: str, pr_classification: dict[str, str]) -> str:
    """Categorise an exercise for the Progression sheet. Use the PR
    sheet's groupings as authority; fall back to keyword heuristics for
    exercises not listed in PRs (variants like 'Incline bench smith',
    'DL romanian', 'Low bar squat smith' still resolve correctly)."""
    if name in pr_classification:
        return pr_classification[name]
    n = name.lower()
    if "squat" in n or "hip thrust" in n:
        return "Squat"
    if "deadlift" in n or re.match(r"^dl(s|\s|sumo|romanian|$)", n):
        return "Deadlift"
    if re.match(r"^bench\b", n) or "incline bench" in n:
        return "Press"
    return "Non compound"
PROG_MATCH_TOL = 0.005  # within 0.5% of running best counts as a match
PROG_PR_FILL = PatternFill("solid", fgColor="C6EFCE")
PROG_PR_FONT = Font(color="006100", bold=True)
PROG_MATCH_FILL = PatternFill("solid", fgColor="FFEB9C")
PROG_MATCH_FONT = Font(color="9C5700", bold=True)
PROG_DROP_FILL = PatternFill("solid", fgColor="FFC7CE")
PROG_DROP_FONT = Font(color="9C0006", bold=True)
PROG_NA_FILL = PatternFill("solid", fgColor="F2F2F2")
PROG_NA_FONT = Font(color="A6A6A6", italic=True)


def _epley_1rm(weight: float, reps: int) -> float:
    return weight * (1 + reps / 30)


def _quarter_key(d: date) -> tuple[int, int]:
    return (d.year, (d.month - 1) // 3 + 1)


def _quarter_label(k: tuple[int, int]) -> str:
    return f"{k[0]} Q{k[1]}"


def _fmt_weight(w: float) -> str:
    return str(int(w)) if w == int(w) else f"{w:g}"


def _write_progression_sheet(
    ws,
    sessions: list[Session],
    parsed_prs: dict[int, list[tuple[str, str, str]]],
) -> None:
    """PR progression by quarter. One row per exercise trained on more
    than PROG_MIN_DAYS distinct days. Each quarter cell shows the best
    set that quarter (by Epley 1RM), colored against the running best
    from all prior quarters: green = new PR, yellow = match, red = drop.
    Matches abbreviate reps to '-'. Rows are grouped using the PR sheet's
    categories and exercise order; exercises with no data in the last
    PROG_INACTIVE_LOOKBACK quarters drop to an 'Inactive' section."""

    per_ex_sets: dict[str, list[tuple[date, float, int]]] = {}
    per_ex_days: dict[str, set[date]] = {}
    for s in sessions:
        for ex in s.exercises:
            key = _normalize_exercise(ex.name)
            had_numeric = False
            for st in ex.sets:
                if st.weight is None or st.reps is None:
                    continue
                if st.weight <= 0 or st.reps <= 0:
                    continue
                per_ex_sets.setdefault(key, []).append((s.date, st.weight, st.reps))
                had_numeric = True
            if had_numeric:
                per_ex_days.setdefault(key, set()).add(s.date)

    qualifying = {
        n for n, days in per_ex_days.items() if len(days) > PROG_MIN_DAYS
    }
    if not qualifying:
        return

    quarters: set[tuple[int, int]] = set()
    for n in qualifying:
        for d, _, _ in per_ex_sets[n]:
            quarters.add(_quarter_key(d))
    quarter_list = sorted(quarters)

    # Best set per (exercise, quarter) by 1RM
    best: dict[tuple[str, tuple[int, int]], tuple[float, int, float]] = {}
    for n in qualifying:
        for d, w, r in per_ex_sets[n]:
            q = _quarter_key(d)
            rm = _epley_1rm(w, r)
            key = (n, q)
            if key not in best or best[key][2] < rm:
                best[key] = (w, r, rm)

    # Derive exercise order and group classification from the PR sheet.
    # Later years override earlier ones for classification; order is
    # first-seen across all years. Normalize names through aliases.
    pr_classification: dict[str, str] = {}
    pr_order: list[str] = []
    seen: set[str] = set()
    for y in sorted(parsed_prs):
        for g, n, _ in parsed_prs[y]:
            nn = _normalize_exercise(n)
            pr_classification[nn] = g
            if nn not in seen:
                seen.add(nn)
                pr_order.append(nn)

    # Bucket qualifying exercises by group, preserving PR order.
    group_to_exs: dict[str, list[str]] = {}
    in_pr = [n for n in pr_order if n in qualifying]
    for n in in_pr:
        group_to_exs.setdefault(pr_classification[n], []).append(n)
    extras = sorted(qualifying - set(in_pr))
    for n in extras:
        g = _progression_group(n, pr_classification)
        group_to_exs.setdefault(g, []).append(n)

    # Active = has a best entry in any of the last PROG_INACTIVE_LOOKBACK quarters.
    recent_quarters = set(quarter_list[-PROG_INACTIVE_LOOKBACK:])

    def is_active(name: str) -> bool:
        return any((name, q) in best for q in recent_quarters)

    ordered_groups = [g for g in GROUP_ORDER if g in group_to_exs] + [
        g for g in group_to_exs if g not in GROUP_ORDER
    ]

    active_sections: list[tuple[str, list[str]]] = []
    inactive_names: list[str] = []
    for g in ordered_groups:
        active = [n for n in group_to_exs[g] if is_active(n)]
        inactive = [n for n in group_to_exs[g] if not is_active(n)]
        if active:
            active_sections.append((g, active))
        inactive_names.extend(inactive)

    headers = ["Exercise", "Days"] + [_quarter_label(q) for q in quarter_list]
    _write_header(ws, headers)
    n_cols = len(headers)

    def write_banner(row: int, label: str) -> None:
        c = ws.cell(row=row, column=1, value=label)
        c.fill = GROUP_FILL
        c.font = GROUP_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
        for col in range(2, n_cols + 1):
            blank = ws.cell(row=row, column=col, value=None)
            blank.fill = GROUP_FILL
            blank.border = BORDER
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=n_cols)

    def write_exercise_row(row: int, name: str) -> None:
        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.border = BORDER
        name_cell.font = Font(bold=True)

        days_cell = ws.cell(row=row, column=2, value=len(per_ex_days[name]))
        days_cell.border = BORDER
        days_cell.alignment = Alignment(horizontal="right")

        running_max = 0.0
        for col_idx, q in enumerate(quarter_list, start=3):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            entry = best.get((name, q))
            if entry is None:
                cell.value = "n/a"
                cell.fill = PROG_NA_FILL
                cell.font = PROG_NA_FONT
                continue
            w, r, rm = entry
            if running_max == 0.0 or rm > running_max * (1 + PROG_MATCH_TOL):
                cell.value = f"{_fmt_weight(w)}x{r}"
                cell.fill = PROG_PR_FILL
                cell.font = PROG_PR_FONT
            elif rm >= running_max * (1 - PROG_MATCH_TOL):
                cell.value = f"{_fmt_weight(w)}x-"
                cell.fill = PROG_MATCH_FILL
                cell.font = PROG_MATCH_FONT
            else:
                cell.value = f"{_fmt_weight(w)}x{r}"
                cell.fill = PROG_DROP_FILL
                cell.font = PROG_DROP_FONT
            if rm > running_max:
                running_max = rm

    row = 2
    for g, names in active_sections:
        write_banner(row, g)
        row += 1
        for n in names:
            write_exercise_row(row, n)
            row += 1

    if inactive_names:
        write_banner(row, "Inactive")
        row += 1
        for n in inactive_names:
            write_exercise_row(row, n)
            row += 1

    ws.freeze_panes = "C2"
    _autosize(ws)


def _classify_group(names: list[str]) -> str:
    joined = " ".join(n.lower() for n in names)
    if "squat" in joined or "thrust" in joined:
        return "Squat"
    if "deadlift" in joined or re.search(r"\bdl", joined):
        return "Deadlift"
    if "bench" in joined or "press" in joined:
        return "Press"
    return "Compound"


def _parse_pr_lines(pr_lines: list[str]) -> list[tuple[str, str, str]]:
    """Parse a PR section into [(group, exercise, sets), ...] preserving
    original order. Groupings come from blank-line-separated blocks in the
    source (e.g. presses, then squats, then deadlifts) plus the explicit
    "(Non compound)" marker. Handles three layouts:
        Exercise            <- name on own line, sets on next
        110x7 ...
        Exercise            <- name with multiple set-lines below
        30x10 ...
        33x8 ...
        Leg press 190x20 190x20 190x20   <- name and sets on same line
    If the same (group, exercise) appears in multiple snapshots, the last
    one wins."""
    blocks: list[list[str]] = [[]]
    block_is_nc: list[bool] = [False]
    is_non_compound = False
    for raw in pr_lines:
        ln = raw.strip()
        if not ln:
            if blocks[-1]:
                blocks.append([])
                block_is_nc.append(is_non_compound)
            continue
        if ln.lower().startswith("(non compound"):
            is_non_compound = True
            if blocks[-1]:
                blocks.append([])
                block_is_nc.append(is_non_compound)
            else:
                # Current block is empty (post-blank-line); retag it.
                block_is_nc[-1] = is_non_compound
            continue
        if PARENS_LINE_RE.match(ln):
            continue
        blocks[-1].append(ln)

    entries: dict[tuple[str, str], str] = {}
    order: list[tuple[str, str]] = []

    for block, is_nc in zip(blocks, block_is_nc):
        if not block:
            continue
        pairs: list[tuple[str, str]] = []
        cur_name: str | None = None
        cur_sets: list[str] = []

        def flush():
            if cur_name is not None:
                pairs.append((cur_name, " ; ".join(s for s in cur_sets if s)))

        for line in block:
            tokens = line.split()
            set_flags = [_is_set_token(t) for t in tokens]
            if tokens and all(set_flags):
                if cur_name is not None:
                    cur_sets.append(line)
                continue
            flush()
            cur_sets = []
            split_idx = next((i for i, f in enumerate(set_flags) if f), None)
            if split_idx is not None and split_idx > 0:
                cur_name = " ".join(tokens[:split_idx])
                cur_sets.append(" ".join(tokens[split_idx:]))
            else:
                cur_name = line
        flush()

        # Drop stray non-exercise lines that leak past the PRs header
        # (e.g. "Met", "Drew ..." at the tail of some logs).
        pairs = [(n, s) for (n, s) in pairs if s]
        if not pairs:
            continue
        group = "Non compound" if is_nc else _classify_group([n for n, _ in pairs])
        for name, sets in pairs:
            key = (group, name)
            if key not in entries:
                order.append(key)
            entries[key] = sets

    return [(g, n, entries[(g, n)]) for (g, n) in order]


def _write_prs_sheet(ws, prs_by_year: dict[int, list[tuple[str, str, str]]]) -> None:
    """Single PR sheet with group header rows. Columns: Exercise, then one
    per year. Groups are emitted in canonical order (Press, Squat,
    Deadlift, Compound, Non compound)."""
    years = sorted(prs_by_year)
    headers = ["Exercise"] + [str(y) for y in years]
    _write_header(ws, headers)

    # Collect exercises by group, preserving first-seen order.
    groups: dict[str, list[str]] = {}
    for y in years:
        for g, n, _ in prs_by_year[y]:
            groups.setdefault(g, [])
            if n not in groups[g]:
                groups[g].append(n)

    ordered = [g for g in GROUP_ORDER if g in groups] + [
        g for g in groups if g not in GROUP_ORDER
    ]

    row = 2
    n_cols = len(headers)
    for g in ordered:
        c = ws.cell(row=row, column=1, value=g)
        c.fill = GROUP_FILL
        c.font = GROUP_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
        for col in range(2, n_cols + 1):
            blank = ws.cell(row=row, column=col, value=None)
            blank.fill = GROUP_FILL
            blank.border = BORDER
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=n_cols)
        row += 1

        for name in groups[g]:
            sets_per_year = []
            for y in years:
                match = next(
                    (s for (gg, nn, s) in prs_by_year[y] if gg == g and nn == name),
                    "",
                )
                sets_per_year.append(match)
            _write_row(ws, row, [name] + sets_per_year)
            row += 1

    _autosize(ws)


def _merge_pr_snapshots(
    snapshots: list[list[tuple[str, str, str]]],
) -> list[tuple[str, str, str]]:
    """Merge PR snapshots (one per log file) for a single year. Preserves
    first-seen order; for duplicate (group, exercise) keys the later
    snapshot wins (most recent PR numbers)."""
    entries: dict[tuple[str, str], str] = {}
    order: list[tuple[str, str]] = []
    for snap in snapshots:
        for g, n, s in snap:
            key = (g, n)
            if key not in entries:
                order.append(key)
            entries[key] = s
    return [(g, n, entries[(g, n)]) for (g, n) in order]


def build_workbook(
    sessions_by_year: dict[int, list[Session]],
    prs_by_year: dict[int, list[list[str]]],
    out: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    years = [y for y in sorted(sessions_by_year)
             if len(sessions_by_year[y]) >= MIN_SESSIONS_PER_YEAR]

    for y in years:
        ws = wb.create_sheet(f"{y} Sets")
        _write_sets_sheet(ws, sessions_by_year[y])
    for y in years:
        ws = wb.create_sheet(f"{y} Sessions")
        _write_sessions_sheet(ws, sessions_by_year[y])

    parsed_prs: dict[int, list[tuple[str, str, str]]] = {}
    for y in prs_by_year:
        if y not in years:
            continue
        per_file = [_parse_pr_lines(lines) for lines in prs_by_year[y]]
        per_file = [s for s in per_file if s]
        if per_file:
            parsed_prs[y] = _merge_pr_snapshots(per_file)
    if parsed_prs:
        _write_prs_sheet(wb.create_sheet("PRs"), parsed_prs)

    all_sessions = [s for y in years for s in sessions_by_year[y]]
    all_sessions.sort(key=lambda s: s.date)
    _write_progression_sheet(
        wb.create_sheet("Progression"), all_sessions, parsed_prs
    )

    wb.save(out)


# --- cli --------------------------------------------------------------------


def _year_from_filename(p: Path) -> int:
    m = FILENAME_YEAR_RE.search(p.name)
    if not m:
        raise ValueError(f"cannot infer year from filename: {p.name}")
    return int(m.group(1))


def main(argv: list[str]) -> int:
    here = Path(__file__).resolve().parent
    if argv[1:]:
        inputs = [here / p for p in argv[1:]]
    else:
        inputs = sorted(here.glob(DEFAULT_GLOB))
    output = here / DEFAULT_OUTPUT

    if not inputs:
        print("no input files found", file=sys.stderr)
        return 1
    missing = [p for p in inputs if not p.exists()]
    if missing:
        print(f"missing inputs: {missing}", file=sys.stderr)
        return 1

    sessions_by_year: dict[int, list[Session]] = {}
    prs_by_year: dict[int, list[list[str]]] = {}
    for p in inputs:
        base_year = _year_from_filename(p)
        sess, prs = parse_log(p.read_text(encoding="utf-8"), base_year)
        for s in sess:
            sessions_by_year.setdefault(s.date.year, []).append(s)
        if prs:
            prs_by_year.setdefault(base_year, []).append(prs)

    for y in sessions_by_year:
        sessions_by_year[y].sort(key=lambda s: s.date)

    build_workbook(sessions_by_year, prs_by_year, output)

    print(f"wrote {output.name}:")
    for y in sorted(sessions_by_year):
        ss = sessions_by_year[y]
        n_sets = sum(len(e.sets) for s in ss for e in s.exercises)
        n_ex = sum(len(s.exercises) for s in ss)
        print(f"  {y}: {len(ss)} sessions, {n_ex} exercises, {n_sets} sets")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
